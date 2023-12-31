import os
import warnings
from datetime import datetime

import openpyxl
import json


def parse_invoice_data(file_path):

    # Suppress the Data Validation warning
    warnings.filterwarnings("ignore", category=UserWarning,
                            message="Data Validation extension is not supported and will be removed")

    workbook = openpyxl.load_workbook(file_path)
    sheets = workbook.sheetnames

    sheets = sheets[:-1]

    all_data = []  # List to store parsed data of all sheets

    # Loop through all sheets
    for sheet in sheets:
        sheet = workbook[sheet]

        # Extract data from specific cells
        data = {}
        data['invoice_number'] = sheet['F1'].value
        data['sender'] = sheet['D4'].value
        data['passport'] = sheet['D7'].value
        # data['dob'] = str(sheet['D8'].value)

        # Extract and convert datetime data
        try:
            dob_datetime = sheet['D8'].value
            if isinstance(dob_datetime, datetime):
                dob_string = dob_datetime.strftime('%d.%m.%Y')
            else:
                dob_string = str(dob_datetime)
            data['dob'] = dob_string
        except Exception as e:
            print('error dob', e)
            raise f'error dob {e}'

        data['weight'] = sheet['H45'].value
        data['phone'] = sheet['D10'].value

        # Extract product data
        product_data = []
        for row_index in range(15, 44):
            product_name = sheet.cell(row=row_index, column=2).value
            product_info = sheet.cell(row=row_index, column=3).value
            product_quantity = sheet.cell(row=row_index, column=6).value
            product_price = sheet.cell(row=row_index, column=7).value

            # Stop parsing when an empty row is encountered
            if not product_name:
                break

            product = {
                'product_name': product_name,
                'product_info': product_info,
                'product_quantity': product_quantity,
                'product_price': product_price
            }
            product_data.append(product)

        data['products'] = product_data

        # Append the data of the current sheet to the list of all_data
        all_data.append(data)

    return all_data


def parse_backet():
    # Replace 'file_path.xlsx' with the path to your Excel file
    parsed_data = parse_invoice_data('exeles\parsing_file.xlsx')

    # Construct the output file path using os.path.join
    output_file_path = os.path.join('src', 'parsed_file.json')

    # Save the parsed data to a JSON file
    with open(output_file_path, 'w', encoding='utf-8') as json_file:
        json.dump(parsed_data, json_file, ensure_ascii=False, indent=4)

    print(f"Parsed data has been saved to '{output_file_path}'.")

    return output_file_path