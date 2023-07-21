import openpyxl


def parse_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    parsed_data = []

    # Starting from the 2nd row until the end of the data
    for row_index in range(3, sheet.max_row + 1):
        code = sheet.cell(row=row_index, column=2).value
        kod = sheet.cell(row=row_index, column=3).value
        product = sheet.cell(row=row_index, column=4).value

        # Check if all three columns are empty
        if not (code or kod or product):
            break  # Stop parsing when the first empty row is encountered

        parsed_data.append({
            "code": code,
            "kod": kod,
            "product": product.lower().strip()
        })

    return parsed_data


# Replace 'file_path.xlsx' with the path to your Excel file
parsed_data = parse_excel('exeles/Product_codes.xlsx')

# Save in JSON format
import json

# Save the parsed data to a JSON file
with open('src\product_codes.json', 'w', encoding='utf-8') as json_file:
    json.dump(parsed_data, json_file, ensure_ascii=False, indent=4)
